#-------------------------------------------------------------------------------
# Name:        get glad alerts # and csv points from online based on input data
# Purpose:
#
# Author:      Ifeoma.Collins
#
# Created:     10/16/2018
#Update: 11/1/2018
# Copyright:   (c) Ifeoma.Collins 2018
# Licence:     <your licence>
#-------------------------------------------------------------------------------

import json
import requests

import glob

import time

import random

import openpyxl
from openpyxl.compat import range

then = time.time() #Time before the operations start

#save results to excel sheet

xls_file = r"U:\icollins\glad\scripts\results\api_results_glad.xlsx"
wb = openpyxl.load_workbook(filename=xls_file)
glad_single = wb['glad']

glad_sum= wb['glad_sum']

glad_xy= wb['xy']



#unique ID field

uid = 'TARGET_FID'


#time period  #leave blank for wohle time period
time_filter = '2018-09-30,2018-10-30'

#result file name
country= "brazil"
theme = "hansen_5kmgrid_matopiba_22_gt1"

# read in our geojson
with open(r'U:\icollins\ptw\soyextent\grid_script_soymatopiba_5km_22_gt1.geojson') as src:
    data = json.load(src)
    print "data opened"
    print "\n"

api_url = r'http://production-api.globalforestwatch.org/v1/glad-alerts'
download_api_url = r'http://production-api.globalforestwatch.org/v1/glad-alerts/download'  #to get lat, long, and juilan day for each date
base_url = 'http://production-api.globalforestwatch.org/'

# iterate over each feature in our feature collection
print "going over feature collection"
print "\n"
output_list = []



for i, feature in enumerate(data['features']):
    #time.sleep(5)  #let API rest a bit
    print "going through features " + str(i)
    print "\n"

    # create geostore first   #NEW
    #to get past error: Geostore or geojson must be set for complex features
    geostore_payload = {'geojson': feature}
    geostore_url = base_url + 'geostore/'

    #OLD
    #r = requests.post(api_url, json={'geojson': feature})

    #NEW
    r = requests.post(geostore_url, json=geostore_payload)
    geostore_id = r.json()['data']['attributes']['hash']

    print r.text
    print "\n"


    #NEW
    param_sum = {'geostore': geostore_id, 'aggregate_values': False, 'period':time_filter}  #for sum sheet
    param_date = {'geostore': geostore_id, 'aggregate_values': True, 'aggregate_by': 'day', 'period':time_filter}   #for indidvidual sheet Julian dates

    #params = {'geostore': geostore_id, 'aggregate_values': False, 'period':"2018-10-01,2018-10-23"}

    #go through sum and date aggregate methods
    agggregate_methods = [param_sum, param_date]

    for agggregate_method in agggregate_methods:
        print "aggregated method: "
        print agggregate_method
        print "\n"
        r = requests.get(api_url, params=agggregate_method)

        print r.text
        print "\n"

        if r.status_code == 500:
                print 'error in query-- see what is going on'

        else:
            print "get request successful no 500"
            response = r.json()['data']['attributes']
            #response = r.json()['data']['attributes']


            # remove extraneous objects from response
            del response['areaHa'], response['downloadUrls']
            #del response['downloadUrls']

            # save response to the geojson
            feature['api_response'] = response

            output_list.append(feature)


        #print r.json()
        print output_list
        print "\n"

        try:
            if agggregate_method == param_sum:
                print "updating TOTAL SUM excel sheet with results"
                number_of_alerts = r.json()['data']['attributes']['value']
                #area = r.json()['data']['attributes']['areaHA']
                print 'there were {} alerts in feature #{}'.format(number_of_alerts, i)

                #update excel sheet sum tab with total number of alerts for whole time by id
                glad_sum["A" + str(i+2)] = feature['properties'][uid]
                glad_sum["B" + str(i+2)] = number_of_alerts
                print "\n"

            if agggregate_method == param_date:

                print "updating JULIAN DAY excel sheet with results"
                #update single sheet
                #year = r.json()['data']['attributes']['year']


                name = feature['properties'][uid]
                text_list = []  #array for lat/long text
                xls_list = [] #array for populating the excel values
                flat_list = [] #array to flatten list


                glad_results = r.json()['data']['attributes']['value']

                #update single_sum
                for k in range (len(glad_results)):  #how to udpate for more than one feature...
                    print k
                    glad_single_row = glad_single.max_row + 1
                    print glad_single_row
                    print glad_results[k]
                    count = glad_results[k]['count']
                    year = glad_results[k]['year']
                    day = glad_results[k]['day']
                    glad_single["C{}".format(glad_single_row)] = name  #id
                    glad_single["A{}".format(glad_single_row)] = year   #year
                    glad_single["B{}".format(glad_single_row)] = day  #julian day
                    glad_single["D{}".format(glad_single_row)] = count   #glad alerts

                    '''
                    glad_single["C" + str(k+2)] = name  #id
                    glad_single["A" + str(k+2)] = year   #year
                    glad_single["B" + str(k+2)] = day  #julian day
                    glad_single["D" + str(k+2)] = count   #glad alerts
                    '''

                #getting lat long added to xlsx for each feature
                #maybe save each response for per id as array and populate new row for array
                #then populate xls with each row
                lat_long_julian = requests.get(download_api_url, params=agggregate_method)  #to store lat/long/julian date for each in the excel sheet...
                #lat_long_julian = lat_long_julian.text
                print "lat_long_julian"
                print lat_long_julian
                print type(lat_long_julian)
                print "\n"

                results = lat_long_julian.text
                print "results"
                print results
                print type(lat_long_julian.text)


                #glad_xy["B{}".format(glad_xy_row)] = results  #id
                print "\n"

                print "splitlines"
                splitlines = results.splitlines()  #to get rid of \r\n new line method in unicode
                print splitlines
                print "\n"

                for splits in splitlines[1:]:
                    print "split"
                    print splits
                    print type(splits)
                    glad_xy_row = glad_xy.max_row + 1
                    glad_xy["A{}".format(glad_xy_row)] = name
                    print "\n"
                    text_split = splits.split(",")
                    print "text split"
                    print text_split
                    print type(text_split)
                    unicode_to_list = list(text_split)
                    print "unicode list"
                    print unicode_to_list
                    print type(unicode_to_list)
                    print "\n"
                    glad_xy["B{}".format(glad_xy_row)] = unicode_to_list[0] #long
                    glad_xy["C{}".format(glad_xy_row)] = unicode_to_list[1] #lat
                    glad_xy["D{}".format(glad_xy_row)] = unicode_to_list[3] #julian day
                    glad_xy["E{}".format(glad_xy_row)] = unicode_to_list[2] #year
                    print "\n"

        except Exception as e:
            #print error
            print "e error"
            print(e)
            print "\n"


def julian_to_date(year, julian_day):

    return datetime.datetime(int(year), 1, 1) + datetime.timedelta(int(julian_day) - 1)


# Save the file as a new xls file...
print "saving results to a new xlsx file"
timestr = time.strftime("%Y%m%d_%H%M%S")
print timestr
wb.save(r"U:\icollins\glad\scripts\results\api_results_" + country + "_" + theme + "_" +  timestr + ".xlsx")

results_api_xls = r"U:\icollins\glad\scripts\results\api_results_" + country + "_" + theme + "_" +  timestr + ".xlsx"


#convert lat/long to points
#may need to convert xls to csv first... or table view
print "converting lat long in xls to points"
in_table = results_api_xls
out_feature_class = r"U:\icollins\glad\scripts\results\points" + country + "_" + theme + ".shp"
x_coords = "longitude"
y_coords = "latitude"
arcpy.management.XYTableToPoint(in_table, out_feature_class,x_coords, y_coords)

print "\n"
now = time.time() #Time after it finished

print("It took: ", now-then, " seconds")
print "done"